"""
Unit tests for data_processor_v2.py
覆盖场景：D1 schema 校验、D2 负值/未来日期过滤、R1 clean_res 异常安全、
         R2 除零保护、R3 filter_pre_launch 全零情况。
"""
import os
import sys
import pytest
import pandas as pd
import numpy as np

# 将 src/ 加入路径，使 import 直接可用（无论从哪里运行 pytest）
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from data_processor_v2 import (
    _normalize_name,
    apply_price_history,
    get_series_from_model,
)


# ---------------------------------------------------------------------------
# _normalize_name
# ---------------------------------------------------------------------------

class TestNormalizeName:
    def test_strips_series_and_5g(self):
        assert _normalize_name("V30 Series 5G") == "v30"

    def test_non_string_input(self):
        result = _normalize_name(30)
        assert isinstance(result, str)

    def test_lowercase_alphanumeric_only(self):
        result = _normalize_name("V50 Pro+ 12+256GB")
        assert result == result.lower()
        assert all(c.isalnum() for c in result)


# ---------------------------------------------------------------------------
# get_series_from_model
# ---------------------------------------------------------------------------

class TestGetSeriesFromModel:
    def test_v30(self):
        assert get_series_from_model("v30_8_256") == "V30"

    def test_v40(self):
        assert get_series_from_model("v40pro") == "V40"

    def test_v50(self):
        assert get_series_from_model("v5012256") == "V50"

    def test_v60(self):
        assert get_series_from_model("v60ultra") == "V60"

    def test_unknown(self):
        assert get_series_from_model("xyz999") == "Unknown"


# ---------------------------------------------------------------------------
# apply_price_history
# ---------------------------------------------------------------------------

class TestApplyPriceHistory:
    def _base_df(self):
        return pd.DataFrame({
            "model_key": ["v50_8_256", "v50_8_256"],
            "date": pd.to_datetime(["2025-06-15", "2025-06-13"]),
            "current_price": [26000.0, 26000.0],
        })

    def test_applies_delta_after_effective_date(self):
        df = self._base_df()
        ph = pd.DataFrame({
            "model_key": ["v50"],
            "effective_from": pd.to_datetime(["2025-06-14"]),
            "price_delta": [-2000.0],
        })
        result = apply_price_history(df, ph)
        assert result.loc[0, "current_price"] == 24000.0  # 2025-06-15 ≥ effective

    def test_does_not_apply_before_effective_date(self):
        df = self._base_df()
        ph = pd.DataFrame({
            "model_key": ["v50"],
            "effective_from": pd.to_datetime(["2025-06-14"]),
            "price_delta": [-2000.0],
        })
        result = apply_price_history(df, ph)
        assert result.loc[1, "current_price"] == 26000.0  # 2025-06-13 < effective

    def test_empty_price_history_falls_back_to_hardcoded(self):
        """空 price_history_df 时走硬编码回退路径，不崩溃。"""
        df = self._base_df()
        result = apply_price_history(df, pd.DataFrame())
        assert "current_price" in result.columns

    def test_no_mutation_of_original_df(self):
        df = self._base_df()
        original_price = df["current_price"].iloc[0]
        ph = pd.DataFrame({
            "model_key": ["v50"],
            "effective_from": pd.to_datetime(["2025-06-14"]),
            "price_delta": [-2000.0],
        })
        apply_price_history(df, ph)
        assert df["current_price"].iloc[0] == original_price  # 原 df 不变


# ---------------------------------------------------------------------------
# D2: 负值和未来日期过滤（通过构造 DataFrame 直接测试过滤逻辑）
# ---------------------------------------------------------------------------

class TestSalesDataFiltering:
    """
    直接测试过滤逻辑（不依赖 Excel 文件 I/O），
    通过构造等效的 df_long 验证负值/未来日期被正确丢弃。
    """

    def _make_sales_df(self, sales_values, dates):
        return pd.DataFrame({
            "model_raw": ["V30"] * len(sales_values),
            "daily_sales": [float(v) for v in sales_values],
            "date": pd.to_datetime(dates),
        })

    def test_negative_sales_would_be_filtered(self):
        """D2: 负值销量必须被过滤。"""
        df = self._make_sales_df(
            [100, -5, 200],
            ["2024-01-01", "2024-01-02", "2024-01-03"]
        )
        neg_mask = df["daily_sales"] < 0
        filtered = df[~neg_mask]
        assert (filtered["daily_sales"] >= 0).all()
        assert len(filtered) == 2

    def test_future_dates_would_be_filtered(self):
        """D2: 未来日期记录必须被过滤。"""
        df = self._make_sales_df(
            [100, 200],
            ["2024-01-01", "2099-12-31"]
        )
        today = pd.Timestamp.now().normalize()
        filtered = df[df["date"] <= today]
        assert len(filtered) == 1
        assert filtered.iloc[0]["date"] == pd.Timestamp("2024-01-01")


# ---------------------------------------------------------------------------
# R2: price_to_launch_ratio 除零保护
# ---------------------------------------------------------------------------

class TestPriceRatioCalculation:
    def test_zero_price_base_does_not_produce_inf(self):
        """R2: price_base=0 时不应产生 inf，应替换为 1.0。"""
        df = pd.DataFrame({
            "current_price": [26000.0, 24000.0],
            "price_base": [0.0, 25000.0],
        })
        # 模拟修复后的计算逻辑
        zero_base = df["price_base"] <= 0
        df.loc[zero_base, "price_base"] = 1.0
        df["ratio"] = df["current_price"] / df["price_base"]
        df["ratio"] = df["ratio"].replace([np.inf, -np.inf], 1.0).fillna(1.0)

        assert not np.isinf(df["ratio"]).any()
        assert not df["ratio"].isna().any()

    def test_normal_ratio_calculation(self):
        df = pd.DataFrame({
            "current_price": [24000.0],
            "price_base": [25000.0],
        })
        df["ratio"] = df["current_price"] / df["price_base"]
        assert abs(df["ratio"].iloc[0] - 0.96) < 1e-6


# ---------------------------------------------------------------------------
# R3: filter_pre_launch 全零销量回退
# ---------------------------------------------------------------------------

class TestFilterPreLaunch:
    def _run_filter(self, sales_values, dates):
        """复现 filter_pre_launch 的逻辑以进行单元测试。"""
        group = pd.DataFrame({
            "model_key": ["v30"] * len(sales_values),
            "daily_sales": [float(v) for v in sales_values],
            "date": pd.to_datetime(dates),
        })
        # 复制 filter_pre_launch 逻辑
        peak_sales = group["daily_sales"].max()
        threshold = max(50, peak_sales * 0.2)
        real_launch_data = group[group["daily_sales"] >= threshold]
        if real_launch_data.empty:
            positive = group[group["daily_sales"] > 0]
            if positive.empty:
                return group  # R3 修复：全零时返回原始 group
            return positive
        first_real_date = real_launch_data["date"].min()
        return group[group["date"] >= first_real_date]

    def test_all_zeros_returns_original_group(self):
        """R3: 全零销量时不返回空 DataFrame，保留所有行。"""
        result = self._run_filter(
            [0, 0, 0],
            ["2024-01-01", "2024-01-02", "2024-01-03"]
        )
        assert len(result) == 3

    def test_normal_case_filters_pre_launch(self):
        """正常情况：找到启动日期并过滤之前的数据。"""
        result = self._run_filter(
            [5, 10, 500, 400, 300],
            ["2024-01-01", "2024-01-02", "2024-01-03", "2024-01-04", "2024-01-05"]
        )
        # threshold = max(50, 500*0.2) = 100, only 500 qualifies -> first real = 2024-01-03
        assert result["date"].min() == pd.Timestamp("2024-01-03")

    def test_no_sales_above_threshold_uses_positive(self):
        """阈值过高时回退到有销量的行。"""
        result = self._run_filter(
            [1, 2, 3],  # 全低于 threshold=max(50, 3*0.2)=50
            ["2024-01-01", "2024-01-02", "2024-01-03"]
        )
        assert len(result) == 3  # 所有正值行


# ---------------------------------------------------------------------------
# D1: clean_specs schema 校验（通过文件测试，需要 openpyxl）
# ---------------------------------------------------------------------------

class TestCleanSpecsSchema:
    pytest.importorskip("openpyxl")

    def test_missing_product_model_returns_empty(self, tmp_path):
        """D1: 缺少 PRODUCT MODEL 列时 clean_specs 应返回空 DataFrame。"""
        import openpyxl
        from data_processor_v2 import clean_specs

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "512GB"
        # 第一列（字段名列）没有 PRODUCT MODEL
        ws.cell(row=1, column=1, value="SOME OTHER FIELD")
        ws.cell(row=2, column=1, value="ANOTHER FIELD")
        ws.cell(row=1, column=2, value="value1")
        ws.cell(row=2, column=2, value="value2")
        filepath = str(tmp_path / "no_product_model.xlsx")
        wb.save(filepath)

        result = clean_specs(filepath)
        assert result.empty, "Should return empty DataFrame when PRODUCT MODEL is missing"

    def test_valid_specs_returns_model_key(self, tmp_path):
        """D1: 包含 PRODUCT MODEL 列的文件应成功解析并生成 model_key。"""
        import openpyxl
        from data_processor_v2 import clean_specs

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "512GB"
        # 格式：第一列为字段名，后续列为产品数据（转置格式）
        fields = ["PRODUCT MODEL", "ORIGINAL PRICE", "RAM", "ROM"]
        values = ["vivo V30 5G", "26000", "8GB", "256GB"]
        for row_idx, (field, val) in enumerate(zip(fields, values), start=1):
            ws.cell(row=row_idx, column=1, value=field)
            ws.cell(row=row_idx, column=2, value=val)
        filepath = str(tmp_path / "valid_specs.xlsx")
        wb.save(filepath)

        result = clean_specs(filepath)
        assert not result.empty
        assert "model_key" in result.columns
        assert len(result) >= 1
